import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

# Load the Excel file
file_path = "bscs.xlsx"
workbook = openpyxl.load_workbook(file_path)

# Define the search values and the fill color
search_values = ["BSCS VI A", "BSCS VIA"]
fill_color = "8DB4E3"  # Hexadecimal representation of (141, 180, 227)

# Define the fill pattern
fill_pattern = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

# Helper function to find the merged range a cell belongs to
def get_merged_range(sheet, cell):
    for merged_cells_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_cells_range:
            return merged_cells_range
    return None

# Iterate over all sheets in the workbook
for sheet in workbook.worksheets:

    # Read the Excel file with pandas, excluding the first two rows
    df = pd.read_excel(file_path, sheet_name=sheet.title, header=None, skiprows=2)

    # Drop the first column
    df.drop(columns=0, inplace=True)

    # Iterate through the DataFrame and find cells containing any of the search values
    for col in range(df.shape[1]):
        for row in range(df.shape[0]):
            cell_value = str(df.iloc[row, col]).replace('\n', ' ').replace('\r', '')  # Remove newline characters

            if any(search_value in cell_value for search_value in search_values):
                cell = sheet.cell(row=row+3, column=col+2)  # Adjust the indices to match the original Excel sheet
                cell.fill = fill_pattern

                # If the cell is part of a merged range, fill all cells in the range
                merged_range = get_merged_range(sheet, cell)
                if merged_range is not None:
                    for merged_row in sheet.iter_rows(min_row=merged_range.min_row,
                                                      max_row=merged_range.max_row,
                                                      min_col=merged_range.min_col,
                                                      max_col=merged_range.max_col):
                        for merged_cell in merged_row:
                            merged_cell.fill = fill_pattern

# Save the modified workbook
workbook.save("bscs-vi-a.xlsx")
