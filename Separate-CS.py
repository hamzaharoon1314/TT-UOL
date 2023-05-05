import openpyxl
import re
from openpyxl.utils import get_column_letter

def find_rows(ws, target_patterns):
    for row in ws.iter_rows():
        for cell in row:
            if cell.coordinate in ws.merged_cells:
                for merged_cell_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_cell_range:
                        cell = ws.cell(merged_cell_range.min_row, merged_cell_range.min_col)
                        break
            cell_value = str(cell.value).strip().lower()
            if any(re.search(target_pattern, cell_value) for target_pattern in target_patterns):
                return cell.row
    return None

def copy_worksheet(src_ws, dest_ws, start_row, end_row):
    for row in src_ws.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            new_cell = dest_ws.cell(row=cell.row - start_row + 1, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.border = cell.border.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.number_format = cell.number_format
                new_cell.protection = cell.protection.copy()
                new_cell.alignment = cell.alignment.copy()

    for merged_cell_range in src_ws.merged_cells.ranges:
        if merged_cell_range.min_row >= start_row and merged_cell_range.max_row <= end_row:
            dest_ws.merge_cells(
                start_row=merged_cell_range.min_row - start_row + 1,
                start_column=merged_cell_range.min_col,
                end_row=merged_cell_range.max_row - start_row + 1,
                end_column=merged_cell_range.max_col,
            )

# Suppress the DeprecationWarning messages
import warnings
warnings.simplefilter("ignore", category=DeprecationWarning)

# Load the Excel workbook
file_path = 'W-23 TIME TABLE UOL SGD.xlsx'
src_wb = openpyxl.load_workbook(file_path)

# Create a new workbook for the filtered data
dest_wb = openpyxl.Workbook()
dest_wb.remove(dest_wb.active)

# Loop through all worksheets in the workbook
for src_ws in src_wb.worksheets:
    # Find the rows with 'CS & IT -' and 'CHEMISTRY -' patterns
    start_row = find_rows(src_ws, [r'cs & it -'])
    end_row = find_rows(src_ws, [r'chemistry -'])

    if start_row is not None and end_row is not None:
        # Create a new worksheet with the same name in the destination workbook
        dest_ws = dest_wb.create_sheet(src_ws.title)

        # Copy the desired rows to the new worksheet, including formatting
        copy_worksheet(src_ws, dest_ws, start_row, end_row)

        # Adjust the row height and column width
        for row in dest_ws.iter_rows():
            for cell in row:
                if cell.value:
                    dest_ws.row_dimensions[cell.row].height = 65
                    dest_ws.column_dimensions[get_column_letter(cell.column)].width = 22
    else:
        print(f"One or both target patterns were not found in the worksheet: {src_ws.title}")

# Save the filtered Excel file
dest_wb.save('bscs.xlsx')
